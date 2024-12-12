import subprocess
import math
import json

import pandas as pd
import openpyxl as xl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from backend.app.constants import *
from backend.app.entities import *



class SheetFormater():
    def __init__(self, data_path: str, schedule_path: str):
        self.wb = xl.Workbook()
        self.ws = self.wb.active

        f = open(data_path, 'r')
        data = json.load(f)
        f.close()

        self.lessons = Lessons(**data)
        self.schedule = pd.read_csv(schedule_path, sep=';')

    def __set_border(self, cell_range, border, overwrite=False):
        def get_border_attr(b):
            return {
                'left': getattr(b, 'left'),
                'right': getattr(b, 'right'),
                'top': getattr(b, 'top'),
                'bottom': getattr(b, 'bottom'),
            }

        for row in self.ws[cell_range]:
            for cell in row:
                if overwrite:
                    cell.border = border
                else:
                    saved_border = get_border_attr(cell.border)
                    new_border = get_border_attr(border)
                    cell.border = Border(
                        left=new_border['left'] if new_border['left'] else saved_border['left'],
                        right=new_border['right'] if new_border['right'] else saved_border['right'],
                        top=new_border['top'] if new_border['top'] else saved_border['top'],
                        bottom=new_border['bottom'] if new_border['bottom'] else saved_border['bottom'],
                    )

    def __get_merge_range(self, row: int, col: int, mode='hor'):
        next_row = row
        next_col = col

        while True:
            if mode == 'hor':
                next_col = col+1
            elif mode == 'ver':
                next_row = row+1

            if self.ws.cell(row, col).value == self.ws.cell(next_row, next_col).value:
                row = next_row
                col = next_col
            else:
                break
        return row, col

    def __merge_cells(self):
        i = 2
        while i < self.ws.max_row:  # merge columns
            j = 4
            while j < self.ws.max_column:
                f_value = self.ws.cell(i, j).value
                s_value = self.ws.cell(i, j+1).value
                if f_value == s_value and \
                        f_value != None:
                    row, col = self.__get_merge_range(i, j, 'hor')
                    self.ws.merge_cells(start_row=i, start_column=j,
                                        end_row=row, end_column=col)
                    i = row
                    j = col
                j += 1
            i += 1

        i = 4
        while i < self.ws.max_column:  # merge rows
            j = 2
            while j < self.ws.max_row:
                f_value = self.ws.cell(j, i).value
                s_value = self.ws.cell(j+1, i).value
                if f_value == s_value and \
                        f_value != None:
                    row, col = self.__get_merge_range(j, i, 'ver')
                    self.ws.merge_cells(start_row=j, start_column=i,
                                        end_row=row, end_column=col)
                    i = col
                    j = row
                j += 1
            i += 1

    def __fill_header(self):
        self.ws.freeze_panes = self.ws['d2']

        self.ws['A1'] = 'День недели'
        self.ws['A1'].alignment = Alignment(text_rotation=90)

        self.ws['b1'] = 'Пара'
        self.ws['b1'].alignment = Alignment(text_rotation=90)

        self.ws['c1'] = 'Время\nзанятий'
        self.ws['c1'].alignment = Alignment(wrap_text=True)

        for i, day in zip(range(2, 43, 8), WEEK_DAYS):
            self.ws.merge_cells(range_string=(f'A{i}:A{i+7}'))
            self.ws[f'a{i}'] = day
            self.ws[f'a{i}'].alignment = Alignment(
                horizontal='center', vertical='center', text_rotation=255)

        for i in range(0, 48):
            self.ws[f'b{i+2}'] = i % 8 + 1
            self.ws[f'c{i+2}'] = LESSON_TIME[i % 8]

        for i in range(0, 16):
            cell = self.ws.cell(row=1, column=i+4)
            cell.font = Font(sz=18)
            cell.value = GROUPS[i]
            cell.alignment = Alignment('center', 'center')

    def __hide_unnecessary_columns(self):
        last_column = xl.utils.cell.column_index_from_string('XFD')
        for idx in range(20, last_column+1):
            self.ws.column_dimensions[xl.utils.get_column_letter(
                idx)].hidden = True

    def __change_sizes(self):
        self.ws.column_dimensions['A'].width = 3.3
        self.ws.column_dimensions['B'].width = 3.3
        self.ws.column_dimensions['C'].width = 12
        for i in range(1, 50):
            self.ws.row_dimensions[i+1].height = 70
        for i in range(ord('D'), ord('S')+1):
            self.ws.column_dimensions[chr(i)].width = 20

    def __fill_schedule(self):
        for i, group in enumerate(GROUPS):
            # TODO: review
            if self.schedule.get(group) is None:
                continue
            
            for j, id in enumerate(self.schedule[group].to_list()):
                if math.isnan(id):
                    continue
                cell = self.ws.cell(row=j+2, column=i+4)
                les_name, les_type, teach_name = self.lessons.get_lesson(
                    int(id))
                text = CellRichText(
                    les_name + '\n',
                    TextBlock(InlineFont(b=True), les_type + '\n'),
                    teach_name,
                )
                cell.value = text
                if les_type == 'ЛК':
                    color = PatternFill(start_color=LEC_COLOR,
                                        end_color=LEC_COLOR, fill_type="solid")
                elif les_type == 'ЛАБ':
                    color = PatternFill(start_color=LAB_COLOR,
                                        end_color=LAB_COLOR, fill_type="solid")
                cell.fill = color
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)

    def format(self):
        self.__change_sizes()
        self.__fill_header()
        self.__fill_schedule()
        self.__merge_cells()
        self.__hide_unnecessary_columns()

        for i in range(1, 50):
            if (i - 1) % 8 == 0:
                self.__set_border(f'A{i}:S{i}',
                                  Border(bottom=Side(style='medium')))
            else:
                self.__set_border(f'A{i}:S{i}',
                                  Border(bottom=Side(style='thin')))

        for c in range(ord('A'), ord('S')+1):
            if c < ord('C'):
                style = 'thin'
            else:
                style = 'medium'
            c = chr(c)
            self.__set_border(f'{c}1:{c}49',
                              Border(right=Side(style=style)))

    def save(self, path):
        self.wb.save(path)


# if __name__ == '__main__':
#     a = SheetFormater('data/data.json', 'data/example-1.csv')
#     a.format()
#     a.save('test.xlsx')
#     subprocess.Popen('open test.xlsx', shell=True)
