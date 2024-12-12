import json
import math
import subprocess

import pandas as pd
import openpyxl as xl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from backend.app.entities import *


# ! deprecated :)

LEC_COLOR = '89c4f5'
LAB_COLOR = 'f7d483'

LESSON_TIME = [
    '9:00 - 10:20',
    '10:30 - 11:50',
    '12:00 - 13:20',
    '13:50 - 15:10',
    '15:20 - 16:40',
    '17:00 - 18:20',
    '18:30 - 19:50',
    '20:00 - 21:20'
]

WEEK_DAYS = [
    'ПОНЕДЕЛЬНИК',
    'ВТОРНИК',
    'СРЕДА',
    'ЧЕТВЕРГ',
    'ПЯТНИЦА',
    'СУББОТА',
]

GROUPS = [
    '1 РФ',
    '2 РФ',
    '3 РФ',
    '4 РФ',
    '8 РФ',
    '3 ФЭ',
    '2 АРИСТ',
    '8 АРИСТ',
    '4 КБ',
    '5 КБ',
    '6 КБ',
    '7 КБ',
    '1 ПИ',
    '5 ПИ',
    '6 ПИ',
    '7 ПИ'

]


f = open('data/data.json', 'r')
data = json.load(f)
f.close()
lessons = Lessons(**data)
schedule_df = pd.read_csv('data/example-1.csv', sep=';')


def set_border(ws, cell_range, border, overwrite=False):
    def get_border_attr(b):
        return {
            'left': getattr(b, 'left'),
            'right': getattr(b, 'right'),
            'top': getattr(b, 'top'),
            'bottom': getattr(b, 'bottom'),
        }

    for row in ws[cell_range]:
        for cell in row:
            if overwrite:
                cell.border = border
            else:
                saved_border = get_border_attr(cell.border)
                new_border = get_border_attr(border)
                cell.border = Border(
                    left=new_border['left'] if new_border['left'] else saved_border['left'],
                    right=new_border['right'] if new_border['right'] else saved_border['right'],
                    top=new_border['top'] if new_border['top'] else saved_border['top'],
                    bottom=new_border['bottom'] if new_border['bottom'] else saved_border['bottom'],
                )


def get_merge_range(ws, row: int, col: int, mode='hor'):
    next_row = row
    next_col = col

    while True:
        if mode == 'hor':
            next_col = col+1
        elif mode == 'ver':
            next_row = row+1

        if ws.cell(row, col).value == ws.cell(next_row, next_col).value:
            row = next_row
            col = next_col
        else:
            break
    return row, col


def merge_cells(ws):
    i = 2
    while i < ws.max_row:  # merge columns
        j = 4
        while j < ws.max_column:
            f_value = ws.cell(i, j).value
            s_value = ws.cell(i, j+1).value
            if f_value == s_value and \
                    f_value != None:
                row, col = get_merge_range(ws, i, j, 'hor')
                ws.merge_cells(start_row=i, start_column=j,
                               end_row=row, end_column=col)
                i = row
                j = col
            j += 1
        i += 1

    i = 4
    while i < ws.max_column:  # merge rows
        j = 2
        while j < ws.max_row:
            f_value = ws.cell(j, i).value
            s_value = ws.cell(j+1, i).value
            if f_value == s_value and \
                    f_value != None:
                row, col = get_merge_range(ws, j, i, 'ver')
                ws.merge_cells(start_row=j, start_column=i,
                               end_row=row, end_column=col)
                i = col
                j = row
            j += 1
        i += 1


wb = xl.Workbook()
ws = wb.active
ws.freeze_panes = ws['d2']

# * FILL HEADER
ws['A1'] = 'День недели'
ws['A1'].alignment = Alignment(text_rotation=90)

ws['b1'] = 'Пара'
ws['b1'].alignment = Alignment(text_rotation=90)

ws['c1'] = 'Время\nзанятий'
ws['c1'].alignment = Alignment(wrap_text=True)

for i, day in zip(range(2, 43, 8), WEEK_DAYS):
    ws.merge_cells(range_string=(f'A{i}:A{i+7}'))
    ws[f'a{i}'] = day
    ws[f'a{i}'].alignment = Alignment(
        horizontal='center', vertical='center', text_rotation=255)

for i in range(0, 48):
    ws[f'b{i+2}'] = i % 8 + 1
    ws[f'c{i+2}'] = LESSON_TIME[i % 8]

for i in range(0, 16):
    cell = ws.cell(row=1, column=i+4)
    cell.font = Font(sz=18)
    cell.value = GROUPS[i]
    cell.alignment = Alignment('center', 'center')


# * HIDE UNNECESSARY COLUMNS
last_column = xl.utils.cell.column_index_from_string('XFD')
for idx in range(20, last_column+1):
    ws.column_dimensions[xl.utils.get_column_letter(idx)].hidden = True

# * CHANGE SIZES
ws.column_dimensions['A'].width = 3.3
ws.column_dimensions['B'].width = 3.3
ws.column_dimensions['C'].width = 12
for i in range(1, 50):
    ws.row_dimensions[i+1].height = 70
for i in range(ord('D'), ord('S')+1):
    ws.column_dimensions[chr(i)].width = 20


# * HANDLE CSV
for i, group in enumerate(GROUPS):
    for j, id in enumerate(schedule_df[group].to_list()):
        if math.isnan(id):
            continue
        cell = ws.cell(row=j+2, column=i+4)
        les_name, les_type, teach_name = lessons.get_lesson(int(id))
        text = CellRichText(
            les_name + '\n',
            TextBlock(InlineFont(b=True), les_type + '\n'),
            teach_name,
        )
        cell.value = text
        if les_type == 'ЛК':
            color = PatternFill(start_color=LEC_COLOR,
                                end_color=LEC_COLOR, fill_type="solid")
        elif les_type == 'ЛАБ':
            color = PatternFill(start_color=LAB_COLOR,
                                end_color=LAB_COLOR, fill_type="solid")
        cell.fill = color
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)


# * MERGE CELLS
merge_cells(ws)

# * SET BORDERS
for i in range(1, 50):
    if (i - 1) % 8 == 0:
        set_border(ws, f'A{i}:S{i}', Border(bottom=Side(style='medium')))
    else:
        set_border(ws, f'A{i}:S{i}', Border(bottom=Side(style='thin')))


for c in range(ord('A'), ord('S')+1):
    if c < ord('C'):
        style = 'thin'
    else:
        style = 'medium'
    c = chr(c)
    set_border(ws, f'{c}1:{c}49', Border(right=Side(style=style)))

wb.save('test.xlsx')
subprocess.Popen('open test.xlsx', shell=True)
